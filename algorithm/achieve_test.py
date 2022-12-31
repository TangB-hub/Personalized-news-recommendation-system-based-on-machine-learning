import os
import openpyxl
import openpyxl as op
import csv_transfrom as tran
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import MultinomialNB

from classifier_test import TextClassifier



data_dir = './CS/'

'''
clf = MultinomialNB(alpha=0.001)
model_path = data_dir + 'NBclassifier.pkl'

classifier = TextClassifier(clf, data_dir + '/fearture_space', model_path)
classifier.validation()
'''



'''
clf = RandomForestClassifier(bootstrap=True, oob_score=True, criterion='gini')
model_path = data_dir + 'mpdels/Radfclassifier.pkl'

classifier = TextClassifier(clf, data_dir + '/fearture_space', model_path)
classifier.validation()
'''


clf = LogisticRegression(C=1000.0)
model_path = data_dir + 'models/LRclassifier.pkl'

classifier = TextClassifier(clf, data_dir + '/fearture_space', model_path)
#classifier.validation()

filename='./content/'
path='./news/sinanews.xlsx'
tran.transfrom(filename,path)
files= os.listdir(filename)

goal=[]
for file in files:
#    print(file)

    f = open(filename+file,'r+',encoding='utf-8')
    text=f.read()
    f.close()
    ret = classifier.predict(text_string=text)
    print(ret)
    goal.extend(ret)


#text_string = '互联网一直在不经意中改变人们的生活和娱乐方式。当电视娱乐和纸介娱乐越来越同质化的时候，人们开始渴望一种新鲜的、更刺激的娱乐方式。于是，来自民间的智慧开始显现，网络恶搞应运而生，并迅速风靡中美这两个互联网最发达的国家。恶搞短片在带给人们无限快感的时候，也招来众多的批评。最新的消息称，国家广电总局将把视频纳入统一监管，引导视频带领中国互联网迈入一个新的时代。'
#ret = classifier.predict(text_string=text_string)
#print(ret)

wb = openpyxl.load_workbook(path)
sheet = wb['Sheet1']

# 在第2列的位置插入1列
sheet.insert_cols(8)
sheet['H1']='label'

num_list = goal

bg = op.load_workbook(path)  # 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下

for i in range(1, len(num_list) + 1):
    sheet.cell(1, 8, 'label')
    sheet.cell(i+1, 8, num_list[i - 1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列

bg.save(path)  # 对文件进行保存

